import openpyxl
import os
import sys

from compare import Segment, Word, Language, LanguageFamily

"""code for reading a structured excel file"""

def readAlignedFile(filename):
	"""return aligned language family object"""
	"""assumes that the file is structured properly"""
	wb = openpyxl.load_workbook(filename, read_only=True)
	meanings = wb.get_sheet_names()
	firstsheet = wb.get_sheet_by_name(meanings[0])
	numlangs = firstsheet.max_row-1
	languagenames = [x[0].value for x in firstsheet['A2':'A{}'.format(numlangs+1)]]
	languages = LanguageFamily()
	llist = []
	for i in range(len(languagenames)):
		llist.append(Language(languagenames[i]))
		for j in range(len(meanings)):
			word = []
			for k in range(3,wb.worksheets[j].max_column):
				word.append(wb.worksheets[j].cell(row=i+2, column=k).value)
			llist[i].addAlignedWord(word, meanings[j])
	for lang in llist:
		languages.addLanguage(lang)
	languages.setMeanings(meanings)
	return languages
	


if __name__=="__main__":
	fam = readAlignedFile(sys.argv[1])
	print(fam)
	l1 = fam.languages[8]
	l2 = fam.languages[9]
	scm = l1.buildCorrespondenceMatrix(l2)
	print("  " + "  ".join(l2.phones))
	for i in range(len(scm)):
		print(l1.phones[i] + str(scm[i]))
	print(l1.correspondenceProbability(l2))
	x = Segment("p")
	print(x)
